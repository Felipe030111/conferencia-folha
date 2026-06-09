from __future__ import annotations

import cgi
import json
import os
import re
import tempfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from html import escape
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
UPLOAD_ROOT = ROOT.parent / ".uploads"
PORT = int(os.environ.get("CONFERENCIA_FOLHA_PORT", "8878"))


LAUNCH_LABELS = {
    "adicional noturno rv": ["adicional noturno"],
    "inf.ben.val.alim.pat": ["informativa benef. vale alimentacao pat", "benef. vale alimentacao"],
    "vale pescados": ["vale compras", "vale pescados"],
    "he 60%": ["hora extra 60", "he 60"],
    "he 70%": ["hora extra 70", "he 70"],
    "he 75%": ["hora extra 75", "he 75"],
    "he 100%": ["hora extra 100", "he 100"],
    "feriado": ["feriado trabalhado 100", "he feriado"],
    "he 60% noturna": ["he 60 noturna", "hora extra 60 noturna"],
    "he 70% noturna": ["he 70 noturna", "hora extra 70 noturna"],
    "he 75% noturna": ["he 75 noturna", "hora extra 75 noturna"],
    "he 100% noturna": ["he 100 noturna", "hora extra 100 noturna"],
    "faltas n/ just. dias": ["faltas", "falta"],
    "faltas n/ just.horas": ["faltas", "falta"],
    "inform. vale transp": ["informativa vale transporte"],
    "vale farmacia": ["vale farmacia"],
    "vale": ["vale"],
    "infor.desp.med.titul": ["desp medicas titular"],
    "infor.desp.med.depen": ["desp medicas depend"],
    "inf.des.odonto titul": ["odonto titular"],
    "infor.desp.odo.depen": ["odonto depen"],
    "multa": ["multa"],
    "ajuda de custos": ["ajuda de custos"],
    "interjornada": ["interjornada"],
}

POINT_EVENT_LABELS = {
    "Feriado trabalhado": ["feriado trabalhado 100", "feriado trabalhado"],
    "Horas Noturnas": ["adicional noturno", "horas noturnas"],
    "HE Feriado": ["feriado trabalhado 100", "he feriado"],
    "Falta em Dias": ["faltas", "falta"],
    "Descontos (-) Faltas": ["faltas", "falta"],
    "Faltas": ["faltas", "falta"],
}


@dataclass
class EmployeeLaunch:
    code: str
    name: str
    values: list[dict[str, Any]]


def strip_accents(value: str) -> str:
    table = str.maketrans(
        "áàãâäéèêëíìîïóòõôöúùûüçÁÀÃÂÄÉÈÊËÍÌÎÏÓÒÕÔÖÚÙÛÜÇ",
        "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC",
    )
    return value.translate(table)


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", strip_accents(str(value or "")).lower()).strip()


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip()
    if not text or text in {"-", "--"}:
        return None
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}.*", text):
        return None
    text = text.replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def money(value: Decimal | None) -> str:
    if value is None:
        return "-"
    text = f"{value:,.2f}"
    return "R$ " + text.replace(",", "X").replace(".", ",").replace("X", ".")


def number_pt(value: Decimal | None) -> str:
    if value is None:
        return "-"
    text = f"{value:,.2f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def br_number_patterns(value: Decimal) -> list[str]:
    base = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    no_group = base.replace(".", "")
    no_dec_if_zero = base[:-3] if base.endswith(",00") else base
    return sorted({base, no_group, no_dec_if_zero}, key=len, reverse=True)


def read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse_launches(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    sheet = next(ws for ws in workbook.worksheets if ws.max_row > 1 and ws.max_column > 1)
    company = sheet.cell(row=1, column=1).value or ""
    period = sheet.cell(row=2, column=1).value or ""
    headers = [sheet.cell(row=3, column=col).value for col in range(1, sheet.max_column + 1)]
    employees: list[EmployeeLaunch] = []

    ignored = {"codigo", "código", "nome", "", "none"}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        if not code or not name:
            continue
        values = []
        for index, raw in enumerate(row[2:], start=2):
            label = headers[index] if index < len(headers) else None
            label_norm = norm(label)
            if label_norm in ignored or label_norm.startswith("qtd colunas") or label_norm.startswith("layout"):
                continue
            amount = parse_decimal(raw)
            if amount is None:
                continue
            values.append({"label": str(label), "amount": amount})
        employees.append(EmployeeLaunch(code=code, name=name, values=values))

    return {"company": str(company), "period": str(period), "employees": employees}


def find_employee_segments(payroll_text: str, employees: list[EmployeeLaunch]) -> dict[str, str]:
    compact = re.sub(r"\s+", " ", payroll_text)
    compact_norm = norm(compact)
    positions: list[tuple[int, str]] = []

    for employee in employees:
        code = str(int(employee.code)) if employee.code.isdigit() else employee.code.lstrip("0")
        first = norm(employee.name).split()[0] if employee.name else ""
        if not code or not first:
            continue
        pattern = re.compile(rf"(?<!\d){re.escape(code)}\s*{re.escape(first)}")
        match = pattern.search(compact_norm)
        if match:
            positions.append((match.start(), employee.code))

    positions.sort()
    segments: dict[str, str] = {}
    for index, (start, code) in enumerate(positions):
        end = positions[index + 1][0] if index + 1 < len(positions) else len(compact)
        segments[code] = compact[start:end]
    return segments


def has_amount(segment: str, amount: Decimal) -> bool:
    return any(pattern in segment for pattern in br_number_patterns(amount))


def has_expected_label(segment: str, label: str) -> bool:
    label_norm = norm(label)
    expected = LAUNCH_LABELS.get(label_norm, [label_norm])
    segment_norm = norm(segment)
    return any(item and item in segment_norm for item in expected)


def extract_payroll_event(segment: str, labels: list[str]) -> dict[str, Decimal | None]:
    segment_norm = norm(segment)
    positions = [segment_norm.find(norm(label)) for label in labels if norm(label) in segment_norm]
    positions = [position for position in positions if position >= 0]
    if not positions:
        return {"reference": None, "amount": None}
    position = min(positions)
    before = segment_norm[max(0, position - 90) : position]
    numbers = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", before)
    parsed = [parse_decimal(number) for number in numbers]
    parsed = [number for number in parsed if number is not None]
    if len(parsed) >= 2:
        return {"reference": parsed[-2], "amount": parsed[-1]}
    if len(parsed) == 1:
        return {"reference": None, "amount": parsed[-1]}
    return {"reference": None, "amount": None}


def compare_launches(launches: dict[str, Any], payroll_text: str) -> list[dict[str, Any]]:
    employees: list[EmployeeLaunch] = launches["employees"]
    segments = find_employee_segments(payroll_text, employees)
    results = []

    for employee in employees:
        segment = segments.get(employee.code, "")
        for item in employee.values:
            amount: Decimal = item["amount"]
            labels = LAUNCH_LABELS.get(norm(item["label"]), [norm(item["label"])])
            amount_found = bool(segment) and has_amount(segment, amount)
            label_found = bool(segment) and has_expected_label(segment, item["label"])
            payroll_event = extract_payroll_event(segment, labels)
            payroll_amount = amount if amount_found else payroll_event["amount"]
            if amount_found and label_found:
                status = "ok"
                message = "Lancamento encontrado na folha."
            elif amount_found:
                status = "warning"
                message = "Valor encontrado, mas a rubrica pode estar com nome diferente."
            elif not segment:
                status = "error"
                message = "Colaborador nao localizado na folha."
            else:
                status = "error"
                message = "Valor nao encontrado no demonstrativo do colaborador."
            results.append(
                {
                    "code": employee.code,
                    "name": employee.name,
                    "label": item["label"],
                    "amount": money(amount),
                    "source_amount": money(amount),
                    "payroll_amount": money(payroll_amount),
                    "status": status,
                    "message": message,
                }
            )
    return results


def parse_loans(portal_text: str) -> list[dict[str, Any]]:
    loans = []
    compact = re.sub(r"\s+", " ", portal_text)
    pattern = re.compile(
        r"(?P<cpf>\d{3}\.\d{3}\.\d{3}-\d{2}).{0,180}?(?P<contract>\d{6,}).{0,80}?R\$\s*(?P<value>[\d\.,]+)"
    )
    for match in pattern.finditer(compact):
        amount = parse_decimal(match.group("value"))
        loans.append({"cpf": match.group("cpf"), "contract": match.group("contract"), "amount": amount})
    return loans


def employee_segments_by_cpf(payroll_text: str) -> dict[str, str]:
    compact = re.sub(r"\s+", " ", payroll_text)
    parts = re.split(r"(?=CPF:\s*\d{3}\.\d{3}\.\d{3}-\d{2})", compact)
    by_cpf = {}
    for part in parts:
        match = re.search(r"CPF:\s*(\d{3}\.\d{3}\.\d{3}-\d{2})", part)
        if match:
            by_cpf[match.group(1)] = part
    return by_cpf


def compare_loans(portal_text: str, payroll_text: str) -> list[dict[str, Any]]:
    loans = parse_loans(portal_text)
    segments = employee_segments_by_cpf(payroll_text)
    grouped: dict[str, dict[str, Any]] = {}
    for loan in loans:
        cpf = loan["cpf"]
        grouped.setdefault(cpf, {"cpf": cpf, "amount": Decimal("0.00"), "contracts": 0})
        if loan["amount"] is not None:
            grouped[cpf]["amount"] += loan["amount"]
        grouped[cpf]["contracts"] += 1

    results = []
    for loan in grouped.values():
        segment = segments.get(loan["cpf"], "")
        amount = loan["amount"].quantize(Decimal("0.01"))
        found = bool(segment) and has_amount(segment, amount)
        label_ok = "credito do trabalhador" in norm(segment)
        payroll_event = extract_payroll_event(segment, ["credito do trabalhador"])
        payroll_amount = amount if found else payroll_event["amount"]
        if found and label_ok:
            status = "ok"
            message = "Total de emprestimos encontrado na folha."
        elif found:
            status = "warning"
            message = "Valor encontrado, mas sem rubrica clara de Credito do Trabalhador."
        elif not segment:
            status = "error"
            message = "CPF do portal nao localizado na folha."
        else:
            status = "error"
            message = "Parcela do portal nao localizada na folha."
        results.append(
            {
                "code": "",
                "name": loan["cpf"],
                "cpf": loan["cpf"],
                "label": f"Credito do trabalhador ({loan['contracts']} contrato(s))",
                "amount": money(amount),
                "source_amount": money(amount),
                "payroll_amount": money(payroll_amount),
                "status": status,
                "message": message,
            }
        )
    return results


def parse_time_to_decimal(value: str) -> Decimal:
    sign = Decimal("-1") if value.startswith("-") else Decimal("1")
    clean = value.replace("-", "")
    hours, minutes = clean.split(":", 1)
    result = Decimal(hours) + (Decimal(minutes) / Decimal("60"))
    return (result * sign).quantize(Decimal("0.01"))


def parse_mirror_events(mirror_text: str) -> list[dict[str, Any]]:
    events = []
    chunks = re.split(r"Página\s+\d+\s+de\s+\d+|Pagina\s+\d+\s+de\s+\d+", mirror_text)
    for chunk in chunks:
        compact = re.sub(r"\s+", " ", chunk).strip()
        if not compact:
            continue
        employee_match = re.search(
            r"Funcion.rio\s+(?P<name>.*?)\s+Departamento.*?Matr.cula\s*(?P<code>\d+)",
            compact,
            re.IGNORECASE,
        )
        if not employee_match:
            continue
        name = employee_match.group("name").strip()
        code = employee_match.group("code").zfill(6)
        summary_end = compact.find("Conforme demonstrativo")
        summary = compact[max(0, summary_end - 900) : summary_end] if summary_end >= 0 else compact[-900:]
        metric_patterns = [
            ("Feriado trabalhado", r"Feriado trabalhado\s+(\d+)"),
            ("Horas Noturnas", r"Horas Noturnas\s+(-?\d{2,3}:\d{2})"),
            ("HE Feriado", r"HE Feriado\s+(-?\d{2,3}:\d{2})"),
            ("Falta em Dias", r"Falta em Dias\s+(\d+)"),
            ("Descontos (-) Faltas", r"Descontos\s+\(-\)\s+Faltas\s+(-?\d{2,3}:\d{2})"),
        ]
        for label, pattern in metric_patterns:
            match = re.search(pattern, summary, re.IGNORECASE)
            if not match:
                continue
            raw = match.group(1)
            value = parse_time_to_decimal(raw) if ":" in raw else Decimal(raw).quantize(Decimal("0.01"))
            events.append(
                {
                    "code": code,
                    "name": name,
                    "label": label,
                    "raw_value": raw,
                    "value": value,
                    "value_kind": "hours" if ":" in raw else "count",
                }
            )
        falta_count = len(re.findall(r"\bFALTA\b", compact, re.IGNORECASE))
        if falta_count:
            events.append(
                {
                    "code": code,
                    "name": name,
                    "label": "Faltas",
                    "raw_value": str(falta_count),
                    "value": Decimal(falta_count).quantize(Decimal("0.01")),
                    "value_kind": "count",
                }
            )
    return events


def compare_mirror(mirror_text: str, payroll_text: str) -> list[dict[str, Any]]:
    mirror_events = parse_mirror_events(mirror_text)
    employees = [EmployeeLaunch(code=event["code"], name=event["name"], values=[]) for event in mirror_events]
    segments = find_employee_segments(payroll_text, employees)
    results = []
    for event in mirror_events:
        segment = segments.get(event["code"], "")
        labels = POINT_EVENT_LABELS.get(event["label"], [event["label"]])
        payroll_event = extract_payroll_event(segment, labels)
        payroll_reference = payroll_event["reference"]
        reference_found = payroll_reference is not None and abs(payroll_reference - event["value"]) <= Decimal("0.02")
        label_found = bool(segment) and any(norm(label) in norm(segment) for label in labels)
        if reference_found and label_found:
            status = "ok"
            message = "Evento de ponto encontrado na folha."
        elif label_found:
            status = "warning"
            message = "Evento localizado, mas a referencia na folha esta diferente."
        elif not segment:
            status = "error"
            message = "Colaborador do espelho nao localizado na folha."
        else:
            status = "error"
            message = "Evento de ponto nao localizado na folha."
        results.append(
            {
                "code": event["code"],
                "name": event["name"],
                "label": event["label"],
                "amount": event["raw_value"],
                "source_amount": event["raw_value"],
                "payroll_amount": number_pt(payroll_reference),
                "payroll_value": money(payroll_event["amount"]),
                "status": status,
                "message": message,
            }
        )
    return results


def parse_liquids(payroll_text: str) -> dict[str, dict[str, Any]]:
    compact = re.sub(r"\s+", " ", payroll_text)
    pattern = re.compile(
        r"(?P<code>\d{1,6})\s*(?P<name>[A-ZÁÀÃÂÉÊÍÓÔÕÚÇ\s]{5,}?)\s+Admissão.*?Líquido\s*-\s*>\s*(?P<liquid>[\d\.,]+)",
        re.DOTALL,
    )
    liquids = {}
    for match in pattern.finditer(compact):
        code = match.group("code").zfill(6)
        value = parse_decimal(match.group("liquid"))
        if value is not None:
            liquids[code] = {"code": code, "name": re.sub(r"\s+", " ", match.group("name")).strip(), "liquid": value}
    return liquids


def compare_liquids(current_text: str, previous_text: str | None) -> list[dict[str, Any]]:
    current = parse_liquids(current_text)
    previous = parse_liquids(previous_text) if previous_text else {}
    rows = []
    for code, item in current.items():
        previous_item = previous.get(code)
        previous_value = previous_item["liquid"] if previous_item else None
        diff = item["liquid"] - previous_value if previous_value is not None else None
        rows.append(
            {
                "code": code,
                "name": item["name"],
                "current": money(item["liquid"]),
                "previous": money(previous_value),
                "difference": money(diff),
                "status": "warning" if diff and abs(diff) >= Decimal("100.00") else "ok",
            }
        )
    return rows


def mirror_summary(mirror_text: str | None) -> dict[str, Any] | None:
    if not mirror_text:
        return None
    employees = len(re.findall(r"\bFuncion.rio\b", mirror_text))
    period = re.search(r"Referente Espelho de ponto\s+(\d{2}/\d{2}/\d{4}\s+até\s+\d{2}/\d{2}/\d{4})", mirror_text)
    return {"employees": employees, "period": period.group(1) if period else "Periodo nao identificado"}


def save_upload(field: Any, suffix: str) -> Path | None:
    if field is None or not getattr(field, "filename", ""):
        return None
    UPLOAD_ROOT.mkdir(exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(field.filename).name)
    target = Path(tempfile.mkdtemp(dir=UPLOAD_ROOT)) / safe_name
    with target.open("wb") as output:
        output.write(field.file.read())
    if suffix and target.suffix.lower() != suffix:
        raise ValueError(f"O arquivo {field.filename} precisa ser {suffix}.")
    return target


def build_report(form: cgi.FieldStorage) -> dict[str, Any]:
    selected_company = form.getfirst("company", "").strip()
    selected_period = form.getfirst("period", "").strip()
    launch_path = save_upload(form["launches"], ".xlsx")
    payroll_path = save_upload(form["payroll"], ".pdf")
    loans_path = save_upload(form["loans"], ".pdf") if "loans" in form else None
    mirror_path = save_upload(form["mirror"], ".pdf") if "mirror" in form else None
    previous_path = save_upload(form["previous_payroll"], ".pdf") if "previous_payroll" in form else None
    if not launch_path or not payroll_path:
        raise ValueError("Envie pelo menos a planilha de lancamentos e a folha de pagamento.")

    launches = parse_launches(launch_path)
    payroll_text = read_pdf_text(payroll_path)
    loans_text = read_pdf_text(loans_path) if loans_path else ""
    mirror_text = read_pdf_text(mirror_path) if mirror_path else ""
    previous_text = read_pdf_text(previous_path) if previous_path else None

    launch_results = compare_launches(launches, payroll_text)
    loan_results = compare_loans(loans_text, payroll_text) if loans_text else []
    mirror_results = compare_mirror(mirror_text, payroll_text) if mirror_text else []
    liquid_results = compare_liquids(payroll_text, previous_text)
    all_results = launch_results + loan_results + mirror_results
    totals = {
        "ok": sum(1 for item in all_results if item["status"] == "ok"),
        "warning": sum(1 for item in all_results if item["status"] == "warning"),
        "error": sum(1 for item in all_results if item["status"] == "error"),
    }

    return {
        "company": selected_company or launches["company"],
        "period": selected_period or launches["period"],
        "file_company": launches["company"],
        "file_period": launches["period"],
        "totals": totals,
        "launches": launch_results,
        "loans": loan_results,
        "mirror_checks": mirror_results,
        "liquids": liquid_results,
        "mirror": mirror_summary(mirror_text),
    }


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def do_POST(self) -> None:
        if self.path != "/api/conferencia":
            self.send_error(404)
            return
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type"),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            report = build_report(form)
            payload = json.dumps(report, default=str).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as exc:
            payload = json.dumps({"error": escape(str(exc))}).encode("utf-8")
            self.send_response(400)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)


def main() -> None:
    server = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    print(f"Plataforma aberta em http://127.0.0.1:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
